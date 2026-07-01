import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import io

# --- LÓGICA DE PROCESAMIENTO ---
def procesar_datos(archivo):
    # Leer el archivo sin cabeceras inicialmente
    df_raw = pd.read_excel(archivo, header=None)
    
    # CASO 1: Formato general donde la primera columna es 'Nombre'
    if (df_raw[0] == 'Nombre').any():
        header_idx = df_raw[df_raw[0] == 'Nombre'].index[0]
        df = df_raw.iloc[header_idx + 1:].copy()
        df = df.iloc[:, :4]
        df.columns = ['Nombre', 'Fecha', 'Semana', 'Hora']

    # CASO 2: Formato individual donde la tabla empieza con 'Fecha'
    elif (df_raw[0] == 'Fecha').any():
        header_idx = df_raw[df_raw[0] == 'Fecha'].index[0]
        
        # Buscar el nombre en las filas de arriba
        nombre_completo = "Desconocido"
        for i in range(header_idx - 1, -1, -1):
            celda = str(df_raw.iloc[i, 0])
            if 'Nombre:' in celda or 'Apellido:' in celda:
                try:
                    # Extraer "Fredy" y "Ccorahua" de la cadena
                    nombre_part = celda.split('Nombre:')[1].split('Apellido:')[0].strip()
                    apellido_part = celda.split('Apellido:')[1].split('ID:')[0].strip()
                    nombre_completo = f"{nombre_part} {apellido_part}"
                except IndexError:
                    nombre_completo = celda # Respaldo por si el formato del texto varía
                break
                
        df = df_raw.iloc[header_idx + 1:].copy()
        df = df.iloc[:, :3] # Tomar Fecha, Semana, Hora
        df.columns = ['Fecha', 'Semana', 'Hora']
        df.insert(0, 'Nombre', nombre_completo) # Crear la columna Nombre al inicio
        
    else:
        raise ValueError("No se reconoció el formato del archivo. Faltan las cabeceras 'Nombre' o 'Fecha'.")

    # Limpiar filas que estén vacías
    df = df.dropna(subset=['Nombre', 'Hora'])
    
    # Convertir formatos de fecha y hora
    df['Fecha'] = pd.to_datetime(df['Fecha']).dt.date
    df['Hora'] = df['Hora'].astype(str).str[:5]
    
    # Agrupar por persona y fecha para sacar el primer y último registro (Ingreso y Salida)
    resumen = df.groupby(['Nombre', 'Fecha'])['Hora'].agg(['min', 'max']).reset_index()
    resumen.columns = ['NOMBRE / APELLIDO', 'Fecha', 'Ingreso', 'Salida']
    
    # Si solo marcó una vez en el día, dejar la salida en blanco
    resumen.loc[resumen['Ingreso'] == resumen['Salida'], 'Salida'] = ""
    
    # Pivotar los datos para el formato de la plantilla
    melted = resumen.melt(id_vars=['NOMBRE / APELLIDO', 'Fecha'], 
                          value_vars=['Ingreso', 'Salida'], 
                          var_name='Salida / Ingreso', value_name='Hora')
    melted['Fecha'] = pd.to_datetime(melted['Fecha']).dt.strftime('%d/%m')
    
    plantilla = melted.pivot(index=['NOMBRE / APELLIDO', 'Salida / Ingreso'], 
                             columns='Fecha', values='Hora').reset_index()
    
    # Invertir el orden (primero Salida, luego Ingreso) y ordenar fechas
    plantilla = plantilla.sort_values(by=['NOMBRE / APELLIDO', 'Salida / Ingreso'], ascending=[True, False])
    
    # Reordenar las columnas de fechas
    cols = list(plantilla.columns[:2]) + sorted(list(plantilla.columns[2:]), reverse=True)
    return plantilla[cols]

# --- LÓGICA DE ESTILOS ---
def aplicar_estilos_excel(df_procesado):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    
    naranja = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    plomo = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    fuente = Font(bold=True, name='Arial', size=10)
    alineacion_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alineacion_c = Alignment(horizontal='center', vertical='center')
    borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Cabeceras
    ws.row_dimensions[1].height = 45
    ws.merge_cells('A1:A2')
    ws['A1'].value = "NOMBRE / APELLIDO"
    ws.merge_cells('B1:B2')
    ws['B1'].value = "Salida / Ingreso"
    
    ultima_col = get_column_letter(len(df_procesado.columns))
    ws.merge_cells(f'C1:{ultima_col}1')
    ws['C1'].value = "Hora de Registro de Entrada y Salida"

    for r in [1, 2]:
        for c in range(1, len(df_procesado.columns) + 1):
            celda = ws.cell(row=r, column=c)
            celda.fill = plomo
            celda.font = fuente
            celda.alignment = alineacion_wrap
            celda.border = borde

    for col_num, fecha in enumerate(df_procesado.columns[2:], 3):
        ws.cell(row=2, column=col_num).value = fecha

    # Datos
    for r_idx, row in enumerate(df_procesado.values, 3):
        tipo = row[1]
        for c_idx, value in enumerate(row, 1):
            celda = ws.cell(row=r_idx, column=c_idx)
            celda.value = value
            celda.border = borde
            celda.alignment = alineacion_wrap if c_idx <= 2 else alineacion_c
            if c_idx <= 2: celda.fill = plomo
            else:
                if tipo == "Salida": celda.fill = naranja
            ws.row_dimensions[r_idx].height = 25

    for i in range(3, ws.max_row + 1, 2):
        ws.merge_cells(start_row=i, start_column=1, end_row=i+1, end_column=1)

    ws.column_dimensions['A'].width = 11.82
    ws.column_dimensions['B'].width = 10.18
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9.9

    wb.save(output)
    return output.getvalue()

# --- INTERFAZ WEB ---
st.set_page_config(page_title="Asistencia Charlotte", page_icon="📅")
st.title("🚀 Transformador de Asistencia")
st.markdown("Sube el archivo de **Transacciones** y obtén la **Plantilla de Directivos** al instante.")

archivo = st.file_uploader("Selecciona el archivo Excel del sistema", type=["xlsx"])

if archivo:
    try:
        with st.spinner('Procesando datos...'):
            df_listo = procesar_datos(archivo)
            excel_final = aplicar_estilos_excel(df_listo)
            
            st.success("✅ ¡Archivo procesado con éxito!")
            st.download_button(
                label="⬇️ Descargar Excel para la Secretaria",
                data=excel_final,
                file_name="Asistencia_Procesada.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"Hubo un error al procesar el archivo: {e}")
